import csv
import re

from django.contrib.auth.decorators import permission_required
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.chart import BarChart
from openpyxl.chart import Reference

import reports.reports.index as index_model
import reports.utils as utils


@permission_required("reports.view_report_index")
def index(request):
    context = {
        "report": index_model.IndexTable(),
    }

    return render(request, "reports/index.jinja", context)


@permission_required("reports.view_report")
def report(request):
    # find the report based on the request
    report_class = utils.get_report_by_slug(request.resolver_match.url_name)

    context = {
        "report": report_class(),
    }

    return render(
        request,
        utils.get_template_by_type(report_class.report_template),
        context,
    )


def export_report_to_csv(request, report_slug, current_tab=None):
    report_class = utils.get_report_by_slug(report_slug)
    report_instance = report_class()

    response = HttpResponse(content_type="text/csv")

    if current_tab:
        response["Content-Disposition"] = (
            f'attachment; filename="{report_slug + "_for_" + current_tab}_report.csv"'
        )
        formatted_current_tab = current_tab.capitalize().replace("_", " ")

        # Define a dictionary to map current_tab values to methods
        tab_mapping = {
            report_instance.tab_name: (
                report_instance.headers(),
                report_instance.rows(),
            ),
            report_instance.tab_name2: (
                report_instance.headers2(),
                report_instance.rows2(),
            ),
            report_instance.tab_name3: (
                report_instance.headers3(),
                report_instance.rows3(),
            ),
            report_instance.tab_name4: (
                report_instance.headers4(),
                report_instance.rows4(),
            ),
        }

        # Use the dictionary to get the methods based on current_tab
        methods = tab_mapping.get(formatted_current_tab)

        if methods:
            headers, rows = methods
        else:
            # Raise an exception if current_tab doesn't match any expected values
            raise ValueError(f"Invalid current_tab value: {formatted_current_tab}")
    else:
        response["Content-Disposition"] = (
            f'attachment; filename="{report_slug}_report.csv"'
        )
        headers = (
            report_instance.headers() if hasattr(report_instance, "headers") else None
        )
        rows = report_instance.rows() if hasattr(report_instance, "rows") else None

    writer = csv.writer(response)

    # Check if the report is a table or a chart
    if hasattr(report_instance, "headers"):
        # For table reports
        writer.writerow([header["text"] for header in headers])
        for row in rows:
            for column in row:
                if str(column["text"]).startswith("<a"):
                    match = re.search(r">(.*?)<", column["text"])
                    if match:
                        column["text"] = match.group(1)
            writer.writerow([column["text"] for column in row])
    else:
        writer.writerow(["Date", "Data"])

        for item in report_instance.data():
            writer.writerow([item["x"], item["y"]])

            # Add an additional row with empty values because Excel needs this for data recognition
            writer.writerow(["", ""])

    return response


def export_report_to_excel(request, report_slug):
    report_class = utils.get_report_by_slug(report_slug)
    report_instance = report_class()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{report_slug}_report.xlsx"'
    )

    workbook = Workbook()
    sheet = workbook.active

    sheet.append(["Date", "Data"])

    for item in report_instance.data():
        sheet.append([item["x"], item["y"]])

        # Add an additional row with empty values because Excel needs this for data recognition
        sheet.append(["", ""])

    chart = BarChart()
    data = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=sheet.max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = report_instance.name
    chart.x_axis.title = "Date"
    chart.y_axis.title = "Data"

    chart.width = 40
    chart.height = 20

    sheet.add_chart(chart, "E5")

    workbook.save(response)

    return response
